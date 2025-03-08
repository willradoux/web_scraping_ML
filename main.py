import requests
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import os
import subprocess
import matplotlib.pyplot as plt
import seaborn as sns

sns.set(style='whitegrid')

def fetch_page(query):
    url = f'https://lista.mercadolivre.com.br/{query.replace(" ", "-")}'
    headers = {'User-Agent': 'Mozilla/5.0'}
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        return BeautifulSoup(response.text, 'html.parser')
    return None

def extract_products_from_page(soup):
    items = soup.select('div.poly-card__content')
    products = []
    for item in items:
        name = item.select_one('a.poly-component__title')
        price = item.select_one('span.andes-money-amount__fraction')
        url = name['href'] if name else None
        products.append({
            'name': name.get_text(strip=True) if name else 'N/A',
            'price': float(price.get_text(strip=True).replace('.', '').replace(',', '.')) if price else 0.0,
            'url': f"https://www.mercadolivre.com.br{url}" if url else 'N/A'
        })
    return products

def save_to_excel(products, query):
    df = pd.DataFrame(products)
    output_dir = 'output'
    os.makedirs(output_dir, exist_ok=True)
    file_name = os.path.join(output_dir, f'mercadolivre_{query.replace(" ", "_")}_products.xlsx')
    df.to_excel(file_name, index=False)
    
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    sheet.title = 'Produtos'
    
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    
    prices = df['price']
    avg_price = prices.mean() if not prices.empty else 0
    
    for row in range(2, len(products) + 2):
        price_cell = sheet[f'B{row}']
        if price_cell.value > avg_price:
            price_cell.fill = red_fill
        else:
            price_cell.fill = green_fill
    
    sheet['D1'] = 'Dashboard'
    sheet['D2'] = f'Média de Preços: R$ {avg_price:.2f}'
    sheet['D3'] = f'Produtos Acima da Média: {(prices > avg_price).sum()}'
    sheet['D4'] = f'Produtos Abaixo da Média: {(prices <= avg_price).sum()}'
    
    for col in sheet.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        sheet.column_dimensions[get_column_letter(col[0].column)].width = max_length + 5
    
    workbook.save(file_name)
    
    if os.name == 'nt':
        os.startfile(file_name)
    else:
        subprocess.run(['open', file_name])

def plot_boxplot(products, query):
    df = pd.DataFrame(products)
    plt.figure(figsize=(8, 6))
    sns.boxplot(x=df['price'])
    plt.title(f'Boxplot de Preços - {query}', fontsize=16)
    plt.xlabel('Preço (R$)', fontsize=14)
    plt.tight_layout()
    plt.show()

def main():
    while True: 
        query = input('Digite o produto que deseja pesquisar: ')
        soup = fetch_page(query)
        
        if soup:
            products = extract_products_from_page(soup)
            
            if products:
                print(f'Itens encontrados: {len(products)}')
                save_to_excel(products, query)
                plot_boxplot(products, query)
            else:
                print('Nenhum produto encontrado.')
        else:
            print('Erro ao buscar a página.')
        
        while True:
            try:
                refazer_busca = int(input('Deseja refazer a busca? 1 para Sim, 2 para Não: '))
                if refazer_busca in [1, 2]:
                    break 
                else:
                    print('Entrada inválida! Digite 1 para Sim ou 2 para Não.')
            except ValueError:
                print('Entrada inválida! Digite um número (1 para Sim ou 2 para Não).')
        
        if refazer_busca == 2:
            break
    
    print("Fim da busca.")

if __name__ == '__main__':
    main()